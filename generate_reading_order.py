#!/usr/bin/env python3
"""
Generate reading_order.json and per-phase JSON files from the Excel source of truth.

Usage:
    python generate_reading_order.py

Reads: X-Men_Reading_Order_final.xlsx
Writes:
    data/reading_order.json              (master — all issues, order 1..N)
    data/phases/reading_order_{era}.json (one per era, local order 1..M)
"""
import json
import re
import sys
from datetime import datetime
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("openpyxl is required: pip install openpyxl")
    sys.exit(1)

# ── Config ──────────────────────────────────────────────────
PROJECT_ROOT = Path(__file__).resolve().parent
EXCEL_PATH = PROJECT_ROOT / "X-Men_Reading_Order_final.xlsx"
DATA_DIR = PROJECT_ROOT / "data"
PHASES_DIR = DATA_DIR / "phases"

# ── Helpers ─────────────────────────────────────────────────

def parse_book(book_str: str):
    """
    Parse "Title #Issue" from the Book column.
    Examples:
        "House of X #1"          → ("House of X", 1)
        "X-Men vol. 5 #1"       → ("X-Men vol. 5", 1)
        "Immortal X-Men Annual #1" → ("Immortal X-Men Annual", 1)
    """
    if not book_str:
        return None, None
    # Match the last #N in the string
    match = re.match(r'^(.+?)\s*#(\d+)\s*$', book_str.strip())
    if match:
        title = match.group(1).strip()
        issue = int(match.group(2))
        return title, issue
    # Fallback: no issue number found
    return book_str.strip(), 1


def extract_year(published):
    """Extract year from the Published column (datetime or string)."""
    if isinstance(published, datetime):
        return str(published.year)
    if isinstance(published, str):
        # Try to extract 4-digit year
        m = re.search(r'(\d{4})', published)
        if m:
            return m.group(1)
    return ""


def safe_filename(era: str) -> str:
    """Convert era name to a safe filename."""
    # Replace problematic chars with underscore, collapse multiples
    safe = re.sub(r'[<>:"/\\|?*]', '_', era)
    safe = re.sub(r'\s+', '_', safe)
    safe = re.sub(r'_+', '_', safe)
    return safe.strip('_')


# ── Main ────────────────────────────────────────────────────

def main():
    if not EXCEL_PATH.exists():
        print(f"ERROR: Excel file not found: {EXCEL_PATH}")
        sys.exit(1)

    print(f"Reading {EXCEL_PATH.name}...")
    wb = openpyxl.load_workbook(str(EXCEL_PATH), data_only=True)
    ws = wb.active

    headers = [c.value for c in ws[1]]
    col = {h: i for i, h in enumerate(headers) if h is not None}

    # Validate required columns
    required = ["Book", "Era"]
    for r in required:
        if r not in col:
            print(f"ERROR: Missing required column '{r}'. Found: {list(col.keys())}")
            sys.exit(1)

    # ── Parse all rows ──────────────────────────────────────
    all_issues = []
    skipped = 0

    for row_idx in range(2, ws.max_row + 1):
        cells = [ws.cell(row=row_idx, column=j + 1).value for j in range(len(headers))]
        book = cells[col["Book"]]
        if not book:
            skipped += 1
            continue

        title, issue = parse_book(str(book))
        if not title:
            skipped += 1
            continue

        era = cells[col.get("Era", -1)] if "Era" in col else ""
        event = cells[col.get("Events/Characters/Universes", -1)] if "Events/Characters/Universes" in col else ""
        published = cells[col.get("Published", -1)] if "Published" in col else ""
        writer = cells[col.get("Writer(s)", -1)] if "Writer(s)" in col else ""
        penciller = cells[col.get("Penciller(s)", -1)] if "Penciller(s)" in col else ""
        main_flag = cells[col.get("Main?", -1)] if "Main?" in col else ""

        year = extract_year(published)
        category = "essencial" if str(main_flag or "").strip().lower() == "yes" else "recomendado"

        all_issues.append({
            "title": title,
            "issue": issue,
            "phase": era or "",
            "event": event or "",
            "year": year,
            "category": category,
            "writer": writer or "",
            "penciller": penciller or "",
        })

    print(f"Parsed {len(all_issues)} issues ({skipped} empty rows skipped)")

    # ── Assign sequential order numbers ─────────────────────
    for i, iss in enumerate(all_issues, start=1):
        iss["order"] = i

    # ── Collect eras ────────────────────────────────────────
    eras_ordered = []
    for iss in all_issues:
        if iss["phase"] and iss["phase"] not in eras_ordered:
            eras_ordered.append(iss["phase"])

    print(f"Eras found: {len(eras_ordered)}")
    for era in eras_ordered:
        count = sum(1 for iss in all_issues if iss["phase"] == era)
        print(f"  {era}: {count} issues")

    # ── Write master reading_order.json ─────────────────────
    DATA_DIR.mkdir(exist_ok=True)
    master = {
        "eras": eras_ordered,
        "total_issues": len(all_issues),
        "issues": all_issues,
    }
    master_path = DATA_DIR / "reading_order.json"
    with open(str(master_path), "w", encoding="utf-8") as f:
        json.dump(master, f, indent=2, ensure_ascii=False)
    print(f"\nWrote {master_path} ({len(all_issues)} issues)")

    # ── Write per-phase JSON files ──────────────────────────
    PHASES_DIR.mkdir(exist_ok=True)

    # Remove old phase files first
    for old in PHASES_DIR.glob("reading_order_*.json"):
        old.unlink()
        print(f"  Removed old: {old.name}")

    for era in eras_ordered:
        era_issues = [iss for iss in all_issues if iss["phase"] == era]
        # Assign local order within phase
        phase_data = []
        for local_order, iss in enumerate(era_issues, start=1):
            entry = dict(iss)
            entry["order"] = local_order
            entry["global_order"] = iss["order"]  # keep reference to master order
            phase_data.append(entry)

        phase_file = {
            "section": era,
            "total_issues": len(phase_data),
            "issues": phase_data,
        }
        fname = f"reading_order_{safe_filename(era)}.json"
        fpath = PHASES_DIR / fname
        with open(str(fpath), "w", encoding="utf-8") as f:
            json.dump(phase_file, f, indent=2, ensure_ascii=False)
        print(f"  Wrote {fname} ({len(phase_data)} issues)")

    print(f"\nDone! {len(all_issues)} total issues across {len(eras_ordered)} eras.")


if __name__ == "__main__":
    main()
